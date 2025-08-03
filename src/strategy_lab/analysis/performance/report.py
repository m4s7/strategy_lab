"""Performance report generation and management."""

import json
import logging
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

import pandas as pd

from ..risk.drawdown import DrawdownAnalysis
from ..risk.metrics import RiskMetrics
from ..trade.statistics import TradeStatistics
from .metrics import PerformanceMetrics

logger = logging.getLogger(__name__)


@dataclass
class PerformanceReport:
    """Comprehensive performance analysis report."""

    metrics: PerformanceMetrics
    risk_metrics: RiskMetrics
    drawdown_analysis: DrawdownAnalysis
    trade_statistics: TradeStatistics
    rolling_metrics: pd.DataFrame
    equity_curve: pd.Series
    trades: pd.DataFrame

    # Additional analysis components (added dynamically)
    advanced_metrics: dict[str, float] | None = None
    time_based_metrics: dict[str, float] | None = None
    seasonality: dict[str, float] | None = None
    market_regimes: pd.DataFrame | None = None
    persistence: dict[str, float] | None = None
    trade_clustering: dict[str, float] | None = None
    trade_patterns: dict[str, float] | None = None
    kelly_sizing: dict[str, float] | None = None

    def to_dict(self) -> dict[str, Any]:
        """Convert report to dictionary."""
        return {
            "metrics": self.metrics.to_dict(),
            "risk_metrics": self.risk_metrics.to_dict(),
            "drawdown_analysis": self.drawdown_analysis.get_summary_stats(),
            "trade_statistics": self.trade_statistics.to_dict(),
            "summary": self.get_summary(),
        }

    def get_summary(self) -> dict[str, Any]:
        """Get executive summary of performance."""
        return {
            "performance": {
                "total_return": f"{self.metrics.total_return:.1%}",
                "annual_return": f"{self.metrics.annualized_return:.1%}",
                "sharpe_ratio": self.metrics.sharpe_ratio,
                "max_drawdown": f"{self.metrics.max_drawdown:.1%}",
                "win_rate": f"{self.metrics.win_rate:.1%}",
            },
            "risk": {
                "volatility": f"{self.risk_metrics.volatility:.1%}",
                "var_95": f"{self.risk_metrics.var_95:.1%}",
                "risk_score": self.risk_metrics.get_risk_score(),
            },
            "trading": {
                "total_trades": self.metrics.total_trades,
                "avg_trade": self.metrics.avg_trade,
                "profit_factor": self.metrics.profit_factor,
            },
        }

    def generate_text_report(self) -> str:
        """Generate formatted text report."""
        lines = []
        lines.append("=" * 80)
        lines.append("PERFORMANCE ANALYSIS REPORT")
        lines.append("=" * 80)
        lines.append(
            f"Period: {self.metrics.start_date.date()} to {self.metrics.end_date.date()}"
        )
        lines.append(f"Trading Days: {self.metrics.trading_days}")
        lines.append("")

        # Performance Section
        lines.append("PERFORMANCE METRICS")
        lines.append("-" * 40)
        lines.append(f"Total Return:        {self.metrics.total_return:>10.1%}")
        lines.append(f"Annualized Return:   {self.metrics.annualized_return:>10.1%}")
        lines.append(f"Sharpe Ratio:        {self.metrics.sharpe_ratio:>10.2f}")
        lines.append(f"Sortino Ratio:       {self.metrics.sortino_ratio:>10.2f}")
        lines.append(f"Calmar Ratio:        {self.metrics.calmar_ratio:>10.2f}")
        lines.append("")

        # Risk Section
        lines.append("RISK METRICS")
        lines.append("-" * 40)
        lines.append(f"Volatility:          {self.risk_metrics.volatility:>10.1%}")
        lines.append(f"Max Drawdown:        {self.metrics.max_drawdown:>10.1%}")
        lines.append(f"VaR (95%):           {self.risk_metrics.var_95:>10.1%}")
        lines.append(f"CVaR (95%):          {self.risk_metrics.cvar_95:>10.1%}")
        lines.append(f"Skewness:            {self.risk_metrics.skewness:>10.2f}")
        lines.append(f"Kurtosis:            {self.risk_metrics.kurtosis:>10.2f}")
        lines.append("")

        # Trading Section
        lines.append("TRADING STATISTICS")
        lines.append("-" * 40)
        lines.append(f"Total Trades:        {self.metrics.total_trades:>10d}")
        lines.append(f"Win Rate:            {self.metrics.win_rate:>10.1%}")
        lines.append(f"Profit Factor:       {self.metrics.profit_factor:>10.2f}")
        lines.append(f"Average Trade:       ${self.metrics.avg_trade:>9.2f}")
        lines.append(f"Average Win:         ${self.metrics.avg_win:>9.2f}")
        lines.append(f"Average Loss:        ${self.metrics.avg_loss:>9.2f}")
        lines.append(f"Win/Loss Ratio:      {self.metrics.win_loss_ratio:>10.2f}")
        lines.append("")

        # Drawdown Section
        lines.append("DRAWDOWN ANALYSIS")
        lines.append("-" * 40)
        dd = self.drawdown_analysis
        lines.append(f"Max Drawdown:        {dd.max_drawdown:>10.1%}")
        lines.append(f"Max Duration:        {dd.max_drawdown_duration:>10d} days")
        lines.append(f"Avg Drawdown:        {dd.avg_drawdown:>10.1%}")
        lines.append(f"Avg Recovery:        {dd.avg_recovery_time:>10d} days")
        lines.append(f"Time Underwater:     {dd.time_underwater_pct:>10.1%}")
        lines.append("")

        # Risk Assessment
        lines.append("RISK ASSESSMENT")
        lines.append("-" * 40)
        risk_score = self.risk_metrics.get_risk_score()
        lines.append(f"Overall Risk Score:  {risk_score:>10.0f}/100")
        lines.append(f"Risk of Ruin:        {self.risk_metrics.risk_of_ruin:>10.1%}")
        lines.append("")

        lines.append("=" * 80)

        return "\n".join(lines)

    def save_to_file(self, filepath: Path, format: str = "json"):
        """Save report to file.

        Args:
            filepath: Output file path
            format: Output format ('json', 'csv', 'html')
        """
        if format == "json":
            with open(filepath, "w") as f:
                json.dump(self.to_dict(), f, indent=2, default=str)

        elif format == "csv":
            # Save key metrics to CSV
            metrics_df = pd.DataFrame([self.metrics.to_dict()])
            metrics_df.to_csv(filepath, index=False)

        elif format == "html":
            # Generate HTML report
            html = self._generate_html_report()
            with open(filepath, "w") as f:
                f.write(html)

        else:
            raise ValueError(f"Unsupported format: {format}")

        logger.info(f"Saved performance report to {filepath}")

    def _generate_html_report(self) -> str:
        """Generate HTML report."""
        html = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <title>Performance Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 40px; }}
                h1 {{ color: #333; }}
                h2 {{ color: #666; }}
                table {{ border-collapse: collapse; width: 100%; margin: 20px 0; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #4472C4; color: white; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .metric {{ font-weight: bold; }}
                .positive {{ color: green; }}
                .negative {{ color: red; }}
            </style>
        </head>
        <body>
            <h1>Performance Analysis Report</h1>
            <p>Period: {self.metrics.start_date.date()} to {self.metrics.end_date.date()}</p>

            <h2>Performance Summary</h2>
            <table>
                <tr>
                    <th>Metric</th>
                    <th>Value</th>
                </tr>
                <tr>
                    <td>Total Return</td>
                    <td class="{'positive' if self.metrics.total_return > 0 else 'negative'}">
                        {self.metrics.total_return:.1%}
                    </td>
                </tr>
                <tr>
                    <td>Annualized Return</td>
                    <td>{self.metrics.annualized_return:.1%}</td>
                </tr>
                <tr>
                    <td>Sharpe Ratio</td>
                    <td>{self.metrics.sharpe_ratio:.2f}</td>
                </tr>
                <tr>
                    <td>Maximum Drawdown</td>
                    <td class="negative">{self.metrics.max_drawdown:.1%}</td>
                </tr>
                <tr>
                    <td>Win Rate</td>
                    <td>{self.metrics.win_rate:.1%}</td>
                </tr>
            </table>

            <h2>Risk Analysis</h2>
            <table>
                <tr>
                    <th>Metric</th>
                    <th>Value</th>
                </tr>
                <tr>
                    <td>Volatility</td>
                    <td>{self.risk_metrics.volatility:.1%}</td>
                </tr>
                <tr>
                    <td>Value at Risk (95%)</td>
                    <td>{self.risk_metrics.var_95:.1%}</td>
                </tr>
                <tr>
                    <td>Risk Score</td>
                    <td>{self.risk_metrics.get_risk_score():.0f}/100</td>
                </tr>
            </table>

            <h2>Trading Statistics</h2>
            <table>
                <tr>
                    <th>Metric</th>
                    <th>Value</th>
                </tr>
                <tr>
                    <td>Total Trades</td>
                    <td>{self.metrics.total_trades}</td>
                </tr>
                <tr>
                    <td>Profit Factor</td>
                    <td>{self.metrics.profit_factor:.2f}</td>
                </tr>
                <tr>
                    <td>Average Trade</td>
                    <td>${self.metrics.avg_trade:.2f}</td>
                </tr>
            </table>

            <p><em>Report generated on {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</em></p>
        </body>
        </html>
        """

        return html

    def export_to_excel(self, filepath: Path):
        """Export comprehensive report to Excel file.

        Args:
            filepath: Excel file path
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font
        except ImportError:
            logger.warning("openpyxl not installed. Cannot export to Excel.")
            return

        wb = Workbook()

        # Summary sheet
        ws_summary = wb.active
        ws_summary.title = "Summary"
        self._write_summary_sheet(ws_summary)

        # Metrics sheet
        ws_metrics = wb.create_sheet("Performance Metrics")
        self._write_metrics_sheet(ws_metrics)

        # Trade Analysis sheet
        ws_trades = wb.create_sheet("Trade Analysis")
        self._write_trade_analysis_sheet(ws_trades)

        # Risk Analysis sheet
        ws_risk = wb.create_sheet("Risk Analysis")
        self._write_risk_analysis_sheet(ws_risk)

        # Trades Data sheet
        ws_data = wb.create_sheet("Trade Data")
        self._write_trades_data_sheet(ws_data)

        # Save file
        wb.save(filepath)
        logger.info(f"Saved Excel report to {filepath}")

    def _write_summary_sheet(self, ws):
        """Write executive summary to worksheet."""
        try:
            from openpyxl.styles import Font
        except ImportError:
            Font = None

        # Title
        ws["A1"] = "Performance Analysis Report"
        if Font:
            ws["A1"].font = Font(size=16, bold=True)

        # Period
        ws["A3"] = "Analysis Period"
        ws["B3"] = f"{self.metrics.start_date.date()} to {self.metrics.end_date.date()}"

        # Key metrics
        row = 5
        metrics_to_show = [
            ("Total Return", f"{self.metrics.total_return:.1%}"),
            ("Annual Return", f"{self.metrics.annualized_return:.1%}"),
            ("Sharpe Ratio", f"{self.metrics.sharpe_ratio:.2f}"),
            ("Max Drawdown", f"{self.metrics.max_drawdown:.1%}"),
            ("Win Rate", f"{self.metrics.win_rate:.1%}"),
            ("Total Trades", str(self.metrics.total_trades)),
            ("Profit Factor", f"{self.metrics.profit_factor:.2f}"),
        ]

        for metric_name, metric_value in metrics_to_show:
            ws[f"A{row}"] = metric_name
            ws[f"B{row}"] = metric_value
            row += 1

    def _write_metrics_sheet(self, ws):
        """Write detailed metrics to worksheet."""
        try:
            from openpyxl.styles import Font
        except ImportError:
            Font = None

        metrics_dict = self.metrics.to_dict()

        row = 1
        ws["A1"] = "Metric"
        ws["B1"] = "Value"
        if Font:
            ws["A1"].font = Font(bold=True)
            ws["B1"].font = Font(bold=True)

        row = 2
        for key, value in metrics_dict.items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            if isinstance(value, float):
                ws[f"B{row}"] = f"{value:.4f}"
            else:
                ws[f"B{row}"] = str(value)
            row += 1

    def _write_trade_analysis_sheet(self, ws):
        """Write trade analysis to worksheet."""
        try:
            from openpyxl.styles import Font
        except ImportError:
            Font = None

        trade_dict = self.trade_statistics.to_dict()

        row = 1
        for category, data in trade_dict.items():
            ws[f"A{row}"] = category.replace("_", " ").title()
            if Font:
                ws[f"A{row}"].font = Font(bold=True)
            row += 1

            if isinstance(data, dict):
                for key, value in data.items():
                    ws[f"B{row}"] = str(key)
                    ws[f"C{row}"] = str(value)
                    row += 1
            row += 1

    def _write_risk_analysis_sheet(self, ws):
        """Write risk analysis to worksheet."""
        try:
            from openpyxl.styles import Font
        except ImportError:
            Font = None

        risk_dict = self.risk_metrics.to_dict()

        row = 1
        ws["A1"] = "Risk Metric"
        ws["B1"] = "Value"
        if Font:
            ws["A1"].font = Font(bold=True)
            ws["B1"].font = Font(bold=True)

        row = 2
        for key, value in risk_dict.items():
            ws[f"A{row}"] = key.replace("_", " ").title()
            if isinstance(value, float):
                ws[f"B{row}"] = f"{value:.4f}"
            else:
                ws[f"B{row}"] = str(value)
            row += 1

    def _write_trades_data_sheet(self, ws):
        """Write trades data to worksheet."""
        try:
            from openpyxl.styles import Font
        except ImportError:
            Font = None

        if self.trades.empty:
            return

        # Write column headers
        for col_idx, col_name in enumerate(self.trades.columns, 1):
            ws.cell(row=1, column=col_idx, value=col_name)
            if Font:
                ws.cell(row=1, column=col_idx).font = Font(bold=True)

        # Write data
        for row_idx, row in enumerate(self.trades.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

    def create_interactive_dashboard(self, output_path: str | None = None) -> str:
        """Create interactive HTML dashboard using Plotly.

        Args:
            output_path: Optional output file path

        Returns:
            HTML string
        """
        try:
            import plotly.graph_objects as go
            from plotly.subplots import make_subplots
        except ImportError:
            logger.warning("Plotly not installed. Cannot create interactive dashboard.")
            return ""

        # Create subplots
        fig = make_subplots(
            rows=3,
            cols=2,
            subplot_titles=(
                "Equity Curve",
                "Monthly Returns",
                "Trade Distribution",
                "Rolling Metrics",
                "Drawdown Analysis",
                "Performance Table",
            ),
            specs=[
                [{"type": "scatter"}, {"type": "heatmap"}],
                [{"type": "histogram"}, {"type": "scatter"}],
                [{"type": "scatter"}, {"type": "table"}],
            ],
            vertical_spacing=0.1,
            horizontal_spacing=0.1,
        )

        # 1. Equity Curve
        fig.add_trace(
            go.Scatter(
                x=self.equity_curve.index,
                y=self.equity_curve.values,
                mode="lines",
                name="Equity",
                line=dict(color="blue", width=2),
            ),
            row=1,
            col=1,
        )

        # 2. Monthly Returns Heatmap
        monthly_returns = self.equity_curve.resample("M").last().pct_change()
        returns_pivot = (
            monthly_returns.groupby(
                [monthly_returns.index.year, monthly_returns.index.month]
            )
            .mean()
            .unstack()
        )

        fig.add_trace(
            go.Heatmap(
                z=returns_pivot.values * 100,
                x=[
                    "Jan",
                    "Feb",
                    "Mar",
                    "Apr",
                    "May",
                    "Jun",
                    "Jul",
                    "Aug",
                    "Sep",
                    "Oct",
                    "Nov",
                    "Dec",
                ],
                y=returns_pivot.index,
                colorscale="RdYlGn",
                zmid=0,
                text=[
                    [f"{val:.1f}%" for val in row] for row in returns_pivot.values * 100
                ],
                texttemplate="%{text}",
                textfont={"size": 10},
            ),
            row=1,
            col=2,
        )

        # 3. Trade Distribution
        if not self.trades.empty and "pnl" in self.trades.columns:
            fig.add_trace(
                go.Histogram(
                    x=self.trades["pnl"],
                    nbinsx=30,
                    name="Trade P&L",
                    marker_color="lightblue",
                ),
                row=2,
                col=1,
            )

        # 4. Rolling Metrics
        if not self.rolling_metrics.empty:
            fig.add_trace(
                go.Scatter(
                    x=self.rolling_metrics.index,
                    y=self.rolling_metrics["sharpe_ratio"],
                    mode="lines",
                    name="Rolling Sharpe",
                    line=dict(color="green"),
                ),
                row=2,
                col=2,
            )

        # 5. Drawdown
        fig.add_trace(
            go.Scatter(
                x=self.drawdown_analysis.underwater_curve.index,
                y=self.drawdown_analysis.underwater_curve.values * 100,
                fill="tozeroy",
                name="Drawdown",
                line=dict(color="red"),
                fillcolor="rgba(255,0,0,0.3)",
            ),
            row=3,
            col=1,
        )

        # 6. Performance Table
        table_data = [
            ["Total Return", f"{self.metrics.total_return:.1%}"],
            ["Annual Return", f"{self.metrics.annualized_return:.1%}"],
            ["Sharpe Ratio", f"{self.metrics.sharpe_ratio:.2f}"],
            ["Max Drawdown", f"{self.metrics.max_drawdown:.1%}"],
            ["Win Rate", f"{self.metrics.win_rate:.1%}"],
            ["Total Trades", str(self.metrics.total_trades)],
        ]

        fig.add_trace(
            go.Table(
                header=dict(values=["Metric", "Value"]),
                cells=dict(values=list(zip(*table_data))),
            ),
            row=3,
            col=2,
        )

        # Update layout
        fig.update_layout(
            title="Strategy Performance Dashboard",
            showlegend=False,
            height=1200,
            width=1600,
        )

        # Generate HTML
        html = fig.to_html(include_plotlyjs="cdn")

        if output_path:
            with open(output_path, "w") as f:
                f.write(html)
            logger.info(f"Saved interactive dashboard to {output_path}")

        return html
